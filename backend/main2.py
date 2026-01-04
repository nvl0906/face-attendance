from fastapi import FastAPI, UploadFile, File, Form, WebSocket, Body, HTTPException, Depends, Query, WebSocketDisconnect, status, UploadFile
from fastapi.middleware.cors import CORSMiddleware
import cv2
import numpy as np
import json
import base64
import io
from PIL import Image
from supabase import create_client, Client
from dotenv import load_dotenv
from insightface.app import FaceAnalysis
from datetime import datetime, timedelta, timezone
import os
import shutil
import asyncio
import pickle
from openpyxl import load_workbook
from fastapi.responses import FileResponse
from fastapi.responses import ORJSONResponse
from scipy.spatial.distance import cosine
from liveness import liveness_check
from export_attendance import export_attendance
from merge_attendance import merge_attendance    
import math
from fastapi.security import OAuth2PasswordBearer
from jose import jwt, JWTError
from scipy.spatial.distance import cdist
import cupy as cp
from babel.dates import format_date, format_datetime
from dateutil.parser import parse
import aiofiles
from mapi import mapilogin, send_sms_simple, get_sms_num
from collections import defaultdict
from collections import Counter
from contextlib import asynccontextmanager
from services.notification_service import (
    NotificationService,
    notify_attendance_marked,
    notify_broadcast,
)
from typing import Optional
from pathlib import Path
from urllib.parse import urlparse
import time
import unicodedata

# Global variables

known_faces = {}
todays_marked = []
todays_emplacement = ""
attendance_records = []
all_members = []
user_profile_map = []
attendance_by_user = defaultdict(list)

BASE_DIR = "known_faces"
os.makedirs(BASE_DIR, exist_ok=True)
PICKLE_FOLDER = "known_faces_embeddings"
os.makedirs(PICKLE_FOLDER, exist_ok=True)

mapitoken = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvbWVzc2FnaW5nLm1hcGkubWdcLyIsImlhdCI6MTc2MjM0NTI0NywiZXhwIjoxNzYyMzQ4ODQ3LCJ1c2VybmFtZSI6Im5hdmFsMDkwNiIsInVzZXJpZCI6NTIzfQ.XcoxBUNLwspvCoJJSDF54hrYCxg8B68yChUmKwHBe3o"

# Load .env
load_dotenv()

# Initialize Supabase
url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")
JWT_SECRET = os.environ.get("JWT_SECRET")
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "ES256")
supabase: Client = create_client(url, key)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")


# Background task function
async def refresh_data_periodically():
    while True:
        try:
            await asyncio.sleep(300) 
            
            # Refresh all data
            await get_username_profile()
            await fetch_all_attendance()
            await get_all_members()
            await load_today_attendance()
            await load_today_attendance_one()
            """
            text = "test"
            formatted = text.replace(". ", ".\n\n")
            notification = notify_broadcast(
                title="Test",
                body=formatted,
                data={"screen": "Présence"}
            )
            
            print(f"Data refreshed at {datetime.now()}")
            """
        except Exception as e:
            print(f"Error refreshing data: {e}")

# Lifespan context manager
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Load initial data (replaces @app.on_event("startup"))
    await get_username_profile()
    await fetch_all_attendance()
    await get_all_members()
    await load_known_faces()
    await load_today_attendance()
    await load_today_attendance_one()
    
    # Start background task
    task = asyncio.create_task(refresh_data_periodically())
    
    yield  # App is running
    
    # Shutdown: Cancel background task (replaces @app.on_event("shutdown"))
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass

# Initialize FastAPI
app = FastAPI(default_response_class=ORJSONResponse,lifespan=lifespan)

# CORS (optional for mobile app testing)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://api.tmiattendance.dpdns.org",
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Load InsightFace model
face_app = FaceAnalysis(providers=['CUDAExecutionProvider'])
face_app.prepare(ctx_id=0, det_size=(640, 640))

def create_access_token(data: dict):
    # no expiration, just encode data with Supabase JWT secret
    return jwt.encode(data, JWT_SECRET, algorithm=JWT_ALGORITHM)

def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return {"userid": payload.get("userid"), "username": payload.get("username"), "is_admin": payload.get("is_admin"), "voice": payload.get("voice"), "profile": payload.get("profile")}
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

async def verify_token(token: str):
    """Decode JWT token and raise if invalid"""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except JWTError:
        return None

def haversine(lat1: float, lon1: float, lat2: float, lon2: float, unit: str = 'meters') -> float:
    try:
        lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])
    except (TypeError, ValueError):
        raise ValueError("All coordinates must be numeric values")
    
    # Validate latitude range (-90 to 90)
    if not (-90 <= lat1 <= 90 and -90 <= lat2 <= 90):
        raise ValueError("Latitude must be between -90 and 90 degrees")
    
    # Validate longitude range (-180 to 180)
    if not (-180 <= lon1 <= 180 and -180 <= lon2 <= 180):
        raise ValueError("Longitude must be between -180 and 180 degrees")
    
    # Handle same location case
    if lat1 == lat2 and lon1 == lon2:
        return 0.0
    
    # WGS-84 mean Earth radius in meters (more accurate than 6371000)
    R = 6371008.8
    
    # Convert inputs to CuPy arrays
    lat1_gpu = cp.asarray(lat1, dtype=cp.float64)
    lon1_gpu = cp.asarray(lon1, dtype=cp.float64)
    lat2_gpu = cp.asarray(lat2, dtype=cp.float64)
    lon2_gpu = cp.asarray(lon2, dtype=cp.float64)
    
    # Convert degrees to radians
    lat1_rad = cp.radians(lat1_gpu)
    lon1_rad = cp.radians(lon1_gpu)
    lat2_rad = cp.radians(lat2_gpu)
    lon2_rad = cp.radians(lon2_gpu)
    
    # Calculate differences
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    # Haversine formula
    # Using the more numerically stable form
    a = cp.sin(dlat / 2.0) ** 2 + cp.cos(lat1_rad) * cp.cos(lat2_rad) * cp.sin(dlon / 2.0) ** 2
    
    # Ensure a is in valid range [0, 1] to avoid numerical errors
    a = cp.clip(a, 0.0, 1.0)
    
    # Calculate angular distance in radians
    c = 2.0 * cp.arctan2(cp.sqrt(a), cp.sqrt(1.0 - a))
    
    # Calculate distance in meters
    distance_meters = float(cp.asnumpy(R * c))
    
    # Convert to requested unit
    unit_conversions = {
        'meters': 1.0,
        'm': 1.0,
        'kilometers': 0.001,
        'km': 0.001,
        'miles': 0.000621371,
        'mi': 0.000621371,
        'feet': 3.28084,
        'ft': 3.28084,
        'nautical_miles': 0.000539957,
        'nm': 0.000539957
    }
    
    if unit.lower() not in unit_conversions:
        raise ValueError(f"Invalid unit '{unit}'. Use: meters, km, miles, feet, nautical_miles")
    
    distance = distance_meters * unit_conversions[unit.lower()]
    
    # Return with appropriate precision
    if unit.lower() in ['meters', 'm']:
        return round(distance, 1)  # 0.1m precision
    elif unit.lower() in ['kilometers', 'km']:
        return round(distance, 3)  # 1m precision when converted back
    elif unit.lower() in ['miles', 'mi']:
        return round(distance, 3)  # ~1.6m precision
    else:
        return round(distance, 2)

def liveness(img, faces, thr=0.40):
    f = faces[0]
    x1, y1, x2, y2 = map(int, f.bbox.astype(int))
    bbox = (x1, y1, x2 - x1, y2 - y1)
    is_real, score = liveness_check(img, bbox, decision_threshold = thr,use_screen_guard=False)

    return is_real

def normalizeword(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    return "".join(
        char for char in normalized
        if unicodedata.category(char) != "Mn"
    )

async def get_member_id(username: str):
    response = await asyncio.to_thread(lambda: supabase.table("members").select("*").eq("username", username).execute().data[0] if username else None)
    if not response:
        raise ValueError("Member not found")
    return response["id"]

async def get_emplacement():
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time())
    end_of_day = start_of_day + timedelta(days=1)

    response = await asyncio.to_thread(lambda: supabase.table("attendance")
        .select("emplacement, timestamp")
        .gte("timestamp", start_of_day.isoformat())
        .lt("timestamp", end_of_day.isoformat())
        .limit(1)
        .execute()
    )

    data = response.data
    if data and len(data) > 0:
        value = data[0]["emplacement"]
    else:
        value = "aucun"
    return value

async def get_all_members():
    global all_members, attendance_records
    response = await asyncio.to_thread(lambda: supabase.table("members").select("*").execute().data)
    all_members = response if response else []
    
    # Build unique sessions: key = (emplacement, date) -> latest timestamp
    unique_sessions = {}
    for row in attendance_records:
        emp = row.get("emplacement")
        ts = row.get("timestamp")
        if not emp or not ts:
            continue
        ts_parsed = parse(ts)
        ts_date = ts_parsed.date()
        key = (emp, ts_date)
        if key not in unique_sessions or ts_parsed > parse(unique_sessions[key]):
            unique_sessions[key] = ts

    # Count unique (emplacement, date) per member (not all records)
    member_sessions = defaultdict(set)
    for row in attendance_records:
        member_id = row.get("member_id")
        emp = row.get("emplacement")
        ts = row.get("timestamp")
        if not member_id or not emp or not ts:
            continue
        ts_date = parse(ts).date()
        member_sessions[member_id].add((emp, ts_date))

    total_sessions = len(unique_sessions)

    # Add attendance_count and absence_count to each member
    for member in all_members:
        member_id = member.get("id")
        # Count unique sessions for this member
        present_count = len(member_sessions.get(member_id, set()))
        
        member['attendance_count'] = present_count
        member['absence_count'] = total_sessions - present_count
        
        # Calculate not_seen
        member_last_attendance = None
        for row in attendance_records:
            if row.get("member_id") == member_id:
                ts_parsed = parse(row.get("timestamp"))
                if not member_last_attendance or ts_parsed > member_last_attendance:
                    member_last_attendance = ts_parsed
        
        if member_last_attendance:
            last_date = member_last_attendance.date()
            missed = sum(1 for d in [parse(ts).date() for ts in unique_sessions.values()] if d > last_date)
            member['not_seen'] = missed
        else:
            member['not_seen'] = total_sessions

async def get_username_profile():
    global user_profile_map
    response = await asyncio.to_thread(lambda: supabase.table("members").select("username, profile").execute().data)
    user_profile_map = response if response else []

async def fetch_all_attendance():
    global attendance_records, attendance_by_user
    batch = 1000
    start = 0
    all_rows = []
    while True:
        response = await asyncio.to_thread(
            lambda: supabase.table("attendance")
            .select("user:member_id(username), member_id, emplacement, timestamp")
            .order("timestamp", desc=True)
            .range(start, start + batch - 1)
            .execute()
        )
        if not response.data:
            break

        all_rows.extend(response.data)

        if len(response.data) < batch:
            break
        start += batch

    # 1️⃣ Track unique (emplacement, date) combinations with their latest timestamp
    unique_emplacement_dates = {}
    
    for row in all_rows:
        emp = row["emplacement"]
        ts = row["timestamp"]
        ts_parsed = parse(ts)
        ts_date = ts_parsed.date()
        
        key = (emp, ts_date)
        
        if key not in unique_emplacement_dates or ts_parsed > parse(unique_emplacement_dates[key]):
            unique_emplacement_dates[key] = ts
    
    # 2️⃣ Group attendance records by user
    user_attendance_map = defaultdict(set)
    
    for row in all_rows:
        member_id = row["member_id"]
        emp = row["emplacement"]
        ts = row["timestamp"]
        ts_date = parse(row["timestamp"]).date()
        user_attendance_map[member_id].add((emp, ts_date))
    
    # 3️⃣ Build final list for each user
    attendance_user = {}
    
    for member_id, user_attended in user_attendance_map.items():
        final_list = []
        
        for (emp, date), latest_ts in unique_emplacement_dates.items():
            ts_parsed = parse(latest_ts)
            formatted_date = format_datetime(ts_parsed, "EEEE d MMMM y", locale="mg_MG")
            
            # Check if this user attended this specific emplacement on this date
            is_present = (emp, date) in user_attended
            attendance_status = "present" if is_present else "absent"
            
            final_list.append({
                "emplacement": emp,
                "timestamp": formatted_date,
                "timestamp_raw": latest_ts,
                "date": date.isoformat(),
                "attendance": attendance_status,
                "_sort_key": ts_parsed
            })
        
        # 4️⃣ Sort by date (most recent first)
        final_list.sort(key=lambda x: x["_sort_key"], reverse=True)
        
        # 5️⃣ Remove the internal sort key
        for item in final_list:
            item.pop("_sort_key", None)
        
        attendance_user[member_id] = final_list
    
    attendance_by_user = attendance_user
    attendance_records = all_rows

async def get_update_attendance(userId, emplacement, timestamp):
    global attendance_by_user

    user_attendance = attendance_by_user[userId]
    
    # Find and update the matching record
    for record in user_attendance:
        if record["emplacement"] == emplacement and record["timestamp"] == timestamp:
            record["attendance"] = "present"
            await asyncio.to_thread(lambda: supabase.table("attendance").insert({
                "member_id": userId,
                "emplacement":  emplacement,
                "timestamp": record["timestamp_raw"]
            }).execute())

    await fetch_all_attendance()
    await get_all_members()
    await load_today_attendance()
    await load_today_attendance_one()
    await load_known_faces()

    return True

async def get_user_attendance(userid: str):
    global attendance_by_user
    return attendance_by_user.get(userid, [])

async def load_today_attendance_one():
    global todays_emplacement
    todays_emplacement = ""
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)

    response = await asyncio.to_thread(lambda: supabase.table("attendance")\
        .select("emplacement")\
        .gte("timestamp", start_of_day.isoformat())\
        .lt("timestamp", end_of_day.isoformat())\
        .limit(1)\
        .execute().data)

    todays_emplacement = response[0]["emplacement"] if response else ""
    # return response if response else []

async def load_today_attendance():
    global todays_marked
    todays_marked = []
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)

    response = await asyncio.to_thread(lambda: supabase.table("attendance")\
        .select("user:member_id(username)")\
        .gte("timestamp", start_of_day.isoformat())\
        .lt("timestamp", end_of_day.isoformat())\
        .execute().data)
    if response:
        for record in response:
            todays_marked.append(record["user"]["username"])
    else:
        todays_marked = []
    # return {record["user"]["username"] for record in response} if response else []

async def get_users():
    response = await asyncio.to_thread(lambda: supabase.table("members").select("username").execute().data)
    return [user["username"] for user in response] if response else []

def normalize(vec):
    return vec / np.linalg.norm(vec)

async def load_known_faces():
    global known_faces
    known_faces.clear()
    users = await get_users()  # fetch user list from Supabase
    image_folder = "known_faces"

    for filename in os.listdir(image_folder):
        name, ext = os.path.splitext(filename)
        if ext.lower() not in ['.jpg', '.png']:
            continue

        img_path = os.path.join(image_folder, filename)
        pkl_path = os.path.join(PICKLE_FOLDER, f"{name}.pkl")

        # Use cached pickle if up-to-date
        if os.path.exists(pkl_path) and os.path.getmtime(pkl_path) >= os.path.getmtime(img_path):
            try:
                with open(pkl_path, "rb") as f:
                    emb = pickle.load(f)
                    known_faces[name] = emb  # already normalized
                    continue
            except Exception as e:
                print(f"⚠️ Error loading pickle for {name}: {e}")

        # Read image
        img = cv2.imread(img_path)
        if img is None:
            continue

        # Add border to avoid missing small faces
        img = cv2.copyMakeBorder(img, 50, 50, 50, 50, cv2.BORDER_CONSTANT, value=[255, 255, 255])

        try:
            faces = face_app.get(img)  # runs on GPU
            if faces:
                emb = normalize(faces[0].embedding)
                known_faces[name] = emb

                # Save to pickle
                with open(pkl_path, "wb") as f:
                    pickle.dump(emb, f)
        except Exception as e:
            print(f"⚠️ Error processing image for {name}: {e}")

def match_face(embedding, threshold=0.59999):
    global known_faces
    embedding = normalize(embedding)  # ensure input is normalized
    best_match = None
    best_distance = float("inf")

    for name, known_emb in known_faces.items():
        dist = cosine(embedding,known_emb)
        if dist < threshold and dist < best_distance:
            best_match = name
            best_distance = dist

    return best_match

async def verify_admin(userId: str, admin: bool):
    user = await asyncio.to_thread(lambda: supabase.table("members").select("is_admin").eq("id", userId).execute().data[0])
    if admin == user["is_admin"]:
        return True
    else:
        return False

@app.post("/v2/register")
async def register_student(
    file: UploadFile,
    name: str = Form(...),
    voice: str = Form(...),
    current_user=Depends(get_current_user)
):
    global known_faces
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    # --- 1. Username already exists check (fast path) ---
    image_folder = "known_faces"
    if os.path.exists(os.path.join(image_folder, f"{name}.jpg")):
        return {"status": "error", "message": f"{name} déjà pris!"}

    # --- 2. Read and decode image (async to avoid blocking event loop) ---
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)
    if img is None:
        return {"status": "error", "message": "Image non valide!"}

    # --- 3. Face detection ---
    faces = face_app.get(img)
    if not faces:
        return {"status": "error", "message": "Aucun visage détecté!"}
    if len(faces) > 1:
        return {"status": "error", "message": "Trop de visages détectés!"}

    # --- 5. Face embedding + duplicate check ---
    new_embedding = faces[0].embedding
    matched_name = match_face(new_embedding)
    if matched_name:
        return {"status": "error", "message": f"Visage déjà existant avec le prénom {matched_name}!"}

    if not matched_name:
        # --- 6. Save face image ---
        save_path = os.path.join(image_folder, f"{name}.jpg")
        async with aiofiles.open(save_path, "wb") as buffer:
            await buffer.write(contents)

        # --- 7. Save embedding pickle (non-blocking thread) ---
        emb_pkl_path = os.path.join(PICKLE_FOLDER, f"{name}.pkl")
        await asyncio.to_thread(
            lambda: pickle.dump(new_embedding, open(emb_pkl_path, "wb"))
        )

        # --- 8. Update in-memory embeddings ---
        known_faces[name] = new_embedding

        # --- 9. Save to DB (offloaded to thread to avoid blocking) ---
        await asyncio.to_thread(
            lambda: supabase.table("members")
                            .insert({"username": name, "voice": voice})
                            .execute()
        )
        # --- 10. Reload known faces and all members ---
        await load_known_faces()
        await get_all_members()
        await fetch_all_attendance()
        await load_today_attendance()
        await load_today_attendance_one()

        return {"status": "success", "message": f"{name} ajouté avec succès!"}

@app.post("/v2/recognize")
async def recognize(
    file: UploadFile = File(...),
    emplacement: str = Form(...),
    latitude: float = Form(...),
    longitude: float = Form(...),
    current_user=Depends(get_current_user)
):
    global todays_marked, todays_emplacement, user_profile_map
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    username = current_user.get("username")
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)

    # --- 1. Validate GPS ---
    gps_data = await asyncio.to_thread(
        lambda: supabase.table("gps").select("latitude,longitude").execute().data
    )
    if not gps_data:
        return {"status": "error", "message": "Aucun GPS activé par l'ADMIN aujourd'hui!"}

    admin_loc = gps_data[0]
    distance = haversine(admin_loc["latitude"], admin_loc["longitude"], latitude, longitude)

    dist_data = await asyncio.to_thread(
        lambda: supabase.table("distance").select("dist").execute().data[0]["dist"]
    )

    if distance > dist_data:
        return {"status": "error", "message": f"+ de {dist_data}m inacceptable: {distance}m!"}

    # --- 2. Decode image ---
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)
    if img is None:
        return {"status": "error", "message": "Image non valide!"}

    # --- 3. Detect faces ---
    faces = face_app.get(img)
    if not faces:
        return {"status": "error", "message": "Aucun visage détecté!"}
    
    """
    # --- 4. Liveness detection ---
    if len(faces) == 1:
        is_real = liveness(img,faces)

        if not is_real:
            return {"status": "error", "message": "Centrez-vous bien pour éviter toute fraude!"}
    """
    
    if todays_emplacement:
        supabase_emplacement = todays_emplacement.strip().lower()
        input_emplacement = emplacement.strip().lower()
        if supabase_emplacement != input_emplacement:
            return {
                "status": "error",
                "message": f"Acceptable: {supabase_emplacement}"
            }

    # --- 6. Match embeddings & record attendance ---
    matches, newly_marked, already_marked = set(), set(), set()
    user_profile = []

    for face in faces:
        emb = face.embedding
        name = match_face(emb)
        if not name:
            continue

        if name not in matches:
            matches.add(name)
            for profile in user_profile_map:
                if profile["username"] == name:
                    item = {}
                    item["username"] = name
                    item["profile"] = profile.get("profile", "")
                    user_profile.append(item)
                    break

            if name not in todays_marked:
                todays_marked.append(name)
                if not todays_emplacement:
                    todays_emplacement = emplacement
                    notification = notify_broadcast(
                        title="TMI",
                        body=f"Vous pouvez effectuer votre présence au lieu {todays_emplacement}",
                        data={"screen": "Présence"}
                    )
                newly_marked.add(name)

                member_id = await get_member_id(name)
                await asyncio.to_thread(
                    lambda: supabase.table("attendance").insert({
                        "member_id": member_id,
                        "emplacement": emplacement
                    }).execute()
                )
            else:
                already_marked.add(name)

    # --- 8. Final response ---
    if matches:
        return {
            "status": "success",
            "matches": list(matches),
            "newly_marked": list(newly_marked),
            "already_marked": list(already_marked),
            "user_profile": user_profile
        }
    else:
        return {"status": "error", "message": "Aucun visage reconnu!"}
  
@app.get("/v2/download")
async def download_excel(current_user=Depends(get_current_user)):
    global attendance_records
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    list_path = "tmi_lisitra.xlsx"
    file_path = "tmi_presence_tracker_pivot_with_emplacement_split.xlsx"

    try:
        # Fetch members from Supabase
        members = await asyncio.to_thread(
            lambda: supabase.table("members")
            .select("username, voice")
            .execute()
            .data
        )

        # Load Excel
        wb = load_workbook(list_path)
        ws = wb.active

        # Clear rows except header
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Existing usernames
        existing_names = {str(cell.value) for cell in ws['B'] if cell.value}

        # Append missing members
        for member in members:
            name = member.get("username")
            voice = member.get("voice")
            if name and name not in existing_names:
                ws.append([voice, name])

        wb.save(list_path)

        # Process attendance async-friendly
        await export_attendance(attendance_records)
        await merge_attendance()

        # Return generated file
        if os.path.exists(file_path):
            return FileResponse(
                path=file_path,
                filename="tmi_presence.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        return {"error": "Fichier non trouvé"}

    except Exception as e:
        print("[Error in /download]", e)
        raise HTTPException(status_code=500, detail="Erreur interne serveur")
    
GPS_ID = "00000000-0000-0000-0000-000000000001"
DIST_ID = "00000000-0000-0000-0000-000000000002"

@app.post("/v2/login")
async def login(
    file: UploadFile = File(...),
):

    # ---- Read and decode image ----
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)

    # ---- Detect faces ----
    faces = face_app.get(img)
    if not faces:
        return {"status": "error", "message": "Aucun visage détecté!"}
    if len(faces) != 1:
        return {"status": "error", "message": "Trop de visages détectés!"}
    
    """
    # --- 4. Liveness detection ---
    is_real = liveness(img,faces)

    if not is_real:
        return {"status": "error", "message": "Centrez-vous bien pour éviter toute fraude!"}
    """
    
    # ---- Face recognition ----
    emb = faces[0].embedding
    name = match_face(emb)
    if not name:
        return {"status": "error", "message": "Merci de vous inscrire auprès de l'ADMIN!"}

    # ---- Supabase lookup ----
    def fetch_user_from_supabase(username: str):
        data = supabase.table("members").select("*").eq("username", username).execute().data
        return data[0] if data else None

    user = await asyncio.to_thread(fetch_user_from_supabase, name)
    if not user:
        return {"status": "error", "message": "Utilisateur introuvable dans la base de données!"}

    # ---- Token & GPS ----
    access_token = create_access_token({
        "userid": user["id"],
        "username": user["username"],
        "voice": user["voice"],
        "is_admin": user["is_admin"],
        "profile": user["profile"]
    })

    if (user["is_admin"]):
        return {
            "status": "successadmin",
            "message": f"Bienvenue ADMIN {name}!",
            "access_token": access_token
        }
    else:
        return {
            "status": "successmember",
            "message": f"Bienvenue {name}!",
            "access_token": access_token
        }

@app.websocket("/ws/v2/recognize")
async def ws_recognize(ws: WebSocket, token: str = Query(...)):
    global todays_marked, todays_emplacement
    # ✅ Token verification
    user = await verify_token(token)
    if not user:
        await ws.close(code=403)
        return

    await ws.accept()

    try:
        while True:
            try:
                # ✅ Receive and parse JSON safely
                message = await ws.receive_text()
                data = json.loads(message)

                image_b64 = data.get("image")
                save_emplacement = data.get("emplacement")
                emplacement = data.get("emplacement", "").strip().lower()

                if not image_b64 or not emplacement:
                    await ws.send_json({"status": "error", "message": "Image ou emplacement non défini!"})
                    continue

                if todays_emplacement:
                    supabase_emplacement = todays_emplacement.strip().lower()
                    if supabase_emplacement != emplacement:
                        await ws.send_json({
                            "status": "error",
                            "message": f"Acceptable: {supabase_emplacement}"
                        })
                        await ws.close()
                        break


                img_data = base64.b64decode(image_b64)
                np_img = np.frombuffer(img_data, np.uint8)
                img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)

                # ✅ Face detection
                faces = face_app.get(img)

                matches, newly_marked_today_view, already_marked_today_view = set(), set(), set()

                tz = timezone(timedelta(hours=3))
                now = datetime.now(tz)

                for face in faces:
                    emb = face.embedding
                    name = match_face(emb)

                    if name:
                        matches.add(name)
                        
                        if name not in todays_marked:
                            todays_marked.append(name)
                            if not todays_emplacement:
                                todays_emplacement = save_emplacement
                            newly_marked_today_view.add(name)
                            member_id = await get_member_id(name)
                            # ✅ Insert asynchronously
                            await asyncio.to_thread(
                                lambda: supabase.table("attendance").insert({
                                    "member_id": member_id,
                                    "emplacement": save_emplacement,
                                }).execute()
                            )
                        elif name in todays_marked:
                            already_marked_today_view.add(name)
                await ws.send_json({
                    "status": "success",
                    "users": list(matches),
                    "newly_marked": list(newly_marked_today_view),
                    "already_marked": list(already_marked_today_view)
                })

            except Exception as e:
                print("[Error] Processing frame:", e)
                await ws.send_json({"status": "error", "message": str(e)})

    except Exception as e:
        print(f"[WS] Error: {e}")

@app.post("/v2/search-user")
async def search_user(payload: dict = Body(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    name = payload.get("name", "").strip()
    if not name or len(name) < 2:
        return {"user": None}
    # Search for user by username (case-insensitive, partial match)
    response = await asyncio.to_thread(lambda: supabase.table("members")\
        .select("id, username, is_admin, voice")\
        .ilike("username", f"%{name}%")\
        .execute().data)
    if response:
        return response
    return 

@app.post("/v2/delete-user")
async def delete_user(payload: dict = Body(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    user_id = payload.get("id")
    user_name = payload.get("name")

    # Remove existing photo from Supabase if exists
    existing_url = await asyncio.to_thread(
        lambda: supabase.table("members").select("profile").eq("username", user_name).execute().data[0].get("profile")
    )
    if existing_url:
        parsed_url = urlparse(existing_url)
        
        # Extract path AFTER the bucket name (images/)
        full_path = parsed_url.path
        bucket_prefix = "/storage/v1/object/public/images/"
        if bucket_prefix in full_path:
            existing_path = full_path.split(bucket_prefix, 1)[1]
            await asyncio.to_thread(
                lambda: supabase.storage
                    .from_("images")
                    .remove([existing_path])
            )

    # Delete from Supabase
    await asyncio.to_thread(lambda: supabase.table("members").delete().eq("id", user_id).execute())
    # Optionally, remove face image and embedding
    image_path = os.path.join("known_faces", f"{user_name}.jpg")
    emb_path = os.path.join(PICKLE_FOLDER, f"{user_name}.pkl")
    for path in [image_path, emb_path]:
        if os.path.exists(path):
            os.remove(path)

    await fetch_all_attendance()
    await get_all_members()
    await load_known_faces()
    await load_today_attendance()
    await load_today_attendance_one()

    return {"status": "success", "message": f"{user_name} supprimé avec succès!"}

@app.post("/v2/update-user")
async def update_user(payload: dict = Body(...), current_user=Depends(get_current_user)):
    global known_faces
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    user_id = payload.get("id")
    new_name = payload.get("name")
    new_voice = payload.get("voice")
    new_admin = payload.get("is_admin")
    res = await asyncio.to_thread(lambda: supabase.table("members").select("*").eq("id", user_id).execute().data[0])

    if new_name == res["username"] and new_voice == res["voice"] and new_admin == res["is_admin"] :
        return {"status": "error", "message": "Aucun changement détecté!"}
    
    # Update in Supabase
    if new_name != res["username"]:
        users = await get_users()
        if new_name in users:
            return {"status": "error", "message": f"{new_name} est déjà pris!"}
        # Rename face image and embedding files
        old_image = os.path.join("known_faces", f"{res['username']}.jpg")
        old_emb = os.path.join(PICKLE_FOLDER, f"{res['username']}.pkl")
        new_image = os.path.join("known_faces", f"{new_name}.jpg")
        new_emb = os.path.join(PICKLE_FOLDER, f"{new_name}.pkl")
        if os.path.exists(old_image):
            os.rename(old_image, new_image)
        if os.path.exists(old_emb):
            os.rename(old_emb, new_emb)
        
        existing_url = res["profile"]
        if existing_url:
            parsed_url = urlparse(existing_url)
            
            # Extract path AFTER the bucket name (images/)
            full_path = parsed_url.path
            bucket_prefix = "/storage/v1/object/public/images/"
            if bucket_prefix in full_path:
                existing_path = full_path.split(bucket_prefix, 1)[1]
                supabase_path = f"profile/{normalizeword(new_name)}_{int(time.time())}.jpg"
                await asyncio.to_thread(
                    lambda: supabase.storage
                        .from_("images")
                        .move(existing_path, supabase_path)
                )

            public_url = await asyncio.to_thread(
                lambda: supabase.storage.from_('images').get_public_url(supabase_path)
            )
            if public_url:
                resp = await asyncio.to_thread(lambda: supabase.table("members").update({"username": new_name}).eq("id", user_id).execute())
                if resp:
                    await asyncio.to_thread(
                        lambda: supabase.table("members").update({"profile": public_url}).eq("username", new_name).execute()
                    )

        # Update memory
        if new_name in known_faces:
            known_faces[new_name] = known_faces.pop(new_name)

    if new_voice != res["voice"]:
        await asyncio.to_thread(lambda: supabase.table("members").update({"voice": new_voice}).eq("id", user_id).execute())

    if new_admin != res["is_admin"]:
        await asyncio.to_thread(lambda: supabase.table("members").update({"is_admin": new_admin}).eq("id", user_id).execute())

    await fetch_all_attendance()
    await get_all_members()
    await load_known_faces()
    await load_today_attendance()
    await load_today_attendance_one()

    return {"status": "success", "message": f"{new_name} mis à jour avec succès!"}

@app.get("/v2/emplacement")
async def get_current_emplacement(current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}
    value = await get_emplacement()
    return {"sup_emplacement": value}

@app.get("/v2/mypresence")
async def user_attendance(current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}
    value = await get_user_attendance(current_user.get("userid"))
    return {"status":"success", "message":"success", "mypresence": value}

@app.get("/v2/allusers")
async def all_users(current_user=Depends(get_current_user)):
    global all_members
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    if all_members:
        value = all_members
        value.sort(key=lambda x: x['username'].lower())
        return {"allusers": value}
    else:
        return {"allusers": []}

@app.post("/v2/userpresence")
async def user_only_attendance(userId = Form(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    value = await get_user_attendance(userId)
    return {"userpresence": value}

@app.post("/v2/updatepresence")
async def update_attendance(userId = Form(...), emplacement = Form(...), timestamp = Form(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    value = await get_update_attendance(userId, emplacement, timestamp)
    if value:
        return {"message": "Succès"}

@app.post("/v2/gps")
async def set_gps_admin(latitude: float = Form(...), longitude: float = Form(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    # Update GPS location for admin
    def upsert_gps():
        supabase.table("gps").upsert({"id": GPS_ID, "latitude": latitude, "longitude": longitude}).execute()

    await asyncio.to_thread(upsert_gps)
    return {"message":"GPS enregistré avec succès"}

@app.post("/v2/update-dist")
async def update_distance(payload: dict = Body(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    dist = payload.get("dist")
    # Update GPS location for admin
    def upsert_distance():
        supabase.table("distance").upsert({"id": DIST_ID, "dist": dist}).execute()

    await asyncio.to_thread(upsert_distance)
    return {"message":f"Distance de {dist}m enregistré  avec succès"}

@app.get("/v2/get-dist")
async def update_distance(current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    # Update GPS location for admin
    def get_distance():
        dist = supabase.table("distance").select({"dist"}).execute().data[0]["dist"]
        return dist

    distance = await asyncio.to_thread(get_distance)
    return {"distance": distance}

@app.post("/v2/register-device")
async def register_device(
    expo_push_token: str = Form(...),
    device_type: str = Form(...),
    device_name: Optional[str] = Form(None),
    current_user=Depends(get_current_user)
):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    """
    Register a device for push notifications
    """
    try:
        member_id = current_user.get("userid")
        result = NotificationService.register_device(
            member_id=member_id,
            expo_push_token=expo_push_token,
            device_type=device_type,
            device_name=device_name
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/v2/unregister-device")
async def unregister_device(
    expo_push_token: str = Form(...),
    current_user=Depends(get_current_user)
):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}
    """
    Unregister a device (on logout)
    """
    try:
        member_id = current_user.get("userid")
        result = NotificationService.unregister_device(member_id, expo_push_token)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/v2/userphoto")
async def get_user_photo(current_user: dict = Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    """Get the current user's photo from known_faces folder"""
    username = current_user["username"]  # Adjust based on your JWT payload structure
    photo_path = Path(f"known_faces/{username}.jpg")
    
    if not photo_path.exists():
        raise HTTPException(status_code=404, detail="Photo not found")
    
    return FileResponse(
        photo_path,
        media_type="image/jpeg",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )

@app.post("/v2/userphoto")
async def update_profile(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    """Upload/update user photo in known_faces folder"""
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    username = current_user["username"]
    photo_path = Path(f"known_faces/{username}.jpg")
    supabase_path = f"profile/{normalizeword(username)}_{int(time.time())}.jpg"

    # Read file content ONCE
    raw_bytes = await file.read()

    # Resize to 100x100 using PIL (force exact size)
    try:
        pil_img = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        pil_img = pil_img.resize((200, 200), Image.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format="JPEG", quality=100)
        file_content = buf.getvalue()
    except Exception as e:
        # If resizing fails, fall back to original bytes but log
        print("⚠️ Resize failed, using original bytes:", e)
        file_content = raw_bytes

    # Save locally - write the same bytes we already read
    with photo_path.open("wb") as buffer:
        buffer.write(file_content)

    # Remove existing photo from Supabase if exists
    existing_url = await asyncio.to_thread(
        lambda: supabase.table("members").select("profile").eq("username", username).execute().data[0].get("profile")
    )
    if existing_url:
        parsed_url = urlparse(existing_url)
        
        # Extract path AFTER the bucket name (images/)
        full_path = parsed_url.path
        bucket_prefix = "/storage/v1/object/public/images/"
        if bucket_prefix in full_path:
            existing_path = full_path.split(bucket_prefix, 1)[1]
            await asyncio.to_thread(
                lambda: supabase.storage
                    .from_("images")
                    .remove([existing_path])
            )

    # Upload to Supabase using the file content (bytes)
    response = await asyncio.to_thread(
        lambda: supabase.storage.from_('images').upload(
            supabase_path, 
            file_content, 
            file_options={"upsert": "false"}
        )
    )
    if response:
        public_url = await asyncio.to_thread(
            lambda: supabase.storage.from_('images').get_public_url(supabase_path)
        )
        if public_url:
            await asyncio.to_thread(
                lambda: supabase.table("members").update({"profile": public_url}).eq("username", username).execute()
            )

    return {"status": "success", "photoUrl": public_url, "message": "Photo de profil mise à jour avec succès"}    

@app.get("/messages")
async def get_messages(limit: int = 50, current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    """Get recent messages from database"""
    try:
        admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
        if not admin_check:
            return {"status":"error", "message":"Veuillez vous-reconnecter svp!"}
        result = supabase.table("messages")\
            .select("*")\
            .order("created_at", desc=True)\
            .limit(limit)\
            .execute()
        
        return {"messages": list(reversed(result.data))}
    except Exception as e:
        return {"error": str(e), "messages": []}

@app.post("/messages")
async def create_message(message: str = Form(...), current_user=Depends(get_current_user)):
    admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
    if not admin_check:
        return {"status":"errorAdmin", "message":"Veuillez vous-reconnecter svp!"}

    """Create a new message (admin only)"""
    user_id = current_user.get("userid")
    try: 
        admin_check = await verify_admin(current_user.get("userid"), current_user.get("is_admin"))
        if not admin_check:
            return {"status":"error", "message":"Veuillez vous-reconnecter svp!"}     
        # Insert message
        result = supabase.table("messages").insert({
            "user_id": user_id,
            "message": message,
        }).execute()
        
        if result.data:
            notification = notify_broadcast(
                title="TMI",
                body=result.data[0]["message"],
                data={"screen": "Message"}
            )
            return result.data[0]
        else:
            raise HTTPException(status_code=500, detail="Failed to create message")
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

"""
@app.on_event("startup")
async def startup_event():
    await fetch_all_attendance()
    await get_all_members()
    await load_known_faces()
    await load_today_attendance()
    await load_today_attendance_one()
   
    message="Test mapi"
    recipient="0343180850"
    res = await mapilogin()
    print("Mapi login response:", res)
    resp = await send_sms_simple(recipient, message, mapitoken)
    print(resp["result"])
    response = await get_sms_num(mapitoken)
    print("SMS credits left:", response)
"""
